#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""BGE 细分结果合并回输出（V225 工作区版）

逻辑（对齐 bge_final_fix.py）：
  1. 读取 _bge_细分/*.csv（不含 remaining）→ 已细分词集合 + 标签映射
  2. 从输出 xlsx 的"通用-其他"行中移除已细分词（保留重复行，全部移除）
  3. 将已细分词按新标签重新写入输出（附加行）
  4. 重新按 100万/50万/标签规则拆分输出文件

用法: python bge_merge.py
"""
import sys, io, os, glob, csv, time
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass
import openpyxl

OUT_DIR = r'D:\DSH-分词工作区\输出'
SUB_DIR = os.path.join(OUT_DIR, '_bge_细分')
DATE = '26-08-14'   # 文件名用连字符（Windows 不允许斜杠）


def load_assigned():
    """读取 BGE 细分结果: assigned[kw] = 完整一级标签"""
    assigned = {}
    for fname in os.listdir(SUB_DIR):
        if 'remaining' in fname or not fname.endswith('.csv'):
            continue
        with open(os.path.join(SUB_DIR, fname), encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split(',', 1)
                if len(parts) == 2 and parts[1].strip() and parts[0].strip() != '标签':
                    assigned[parts[1].strip()] = parts[0].strip()
    print(f'已细分词: {len(assigned)} 条')
    return assigned


def main():
    assigned = load_assigned()
    if not assigned:
        print('❌ 无细分结果')
        return

    # 1) 从输出文件读取全部行，过滤通用-其他中的已细分词
    files = sorted(glob.glob(os.path.join(OUT_DIR, '*.xlsx')))
    rows = []          # (一级标签, 二级标签, 关键词)
    moved = 0
    for fp in files:
        wb = openpyxl.load_workbook(fp, read_only=True)
        for ws in wb.worksheets:
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not row or len(row) < 3 or not row[2]:
                    continue
                label, l2, kw = str(row[0]), str(row[1]), str(row[2]).strip()
                if label.startswith('通用-其他-其他-') and kw in assigned:
                    moved += 1   # 移除，稍后按新标签写回
                    continue
                rows.append((label, l2, kw))
        wb.close()
    print(f'读取 {len(rows)+moved} 行，其中通用-其他移除 {moved} 条（将按 BGE 标签写回）')

    # 2) 按 BGE 标签写回
    for kw, label in assigned.items():
        l2 = label.split('-')[-1] if len(label.split('-')) > 1 else '其他'
        # 单元号由最终重新分类时处理，这里简化：L2 不带单元号（保持与原输出一致）
        rows.append((f'{label}-{DATE}', l2, kw))

    # 3) 重新按规则拆分输出（复用 split_output 逻辑）
    import pandas as pd
    df = pd.DataFrame(rows, columns=['一级标签', '二级标签', '关键词'])
    df['L1'] = df['一级标签'].map(lambda x: '行业-AI' if str(x).startswith('行业-AI') else str(x).split('-')[0])
    from collections import Counter
    cnt = Counter(df['L1'])
    print('拆分后 L1 分布:')
    for k, v in cnt.most_common():
        print(f'  {v:>9,} | {k}')

    # 写文件（100万/文件，50万/sheet，>50万单独）
    def write_workbook(out_path, part_rows, sheet_prefix):
        wb = openpyxl.Workbook(write_only=True)
        n_sheets = (len(part_rows) + 500000 - 1) // 500000
        for si in range(n_sheets):
            ws = wb.create_sheet(f'{sheet_prefix}-{si+1}')
            ws.append(['一级标签', '二级标签', '关键词'])
            chunk = part_rows[si*500000:(si+1)*500000]
            for r in chunk:
                ws.append(r)
        wb.save(out_path)
        wb.close()
        return n_sheets

    # 清空旧输出 xlsx（保留 _bge_细分 和说明）
    for fp in files:
        os.remove(fp)

    out_files = []
    solo = [k for k, v in cnt.items() if v > 500000]
    merge = [k for k, v in cnt.items() if v <= 500000]
    for l1 in solo:
        sub = df[df['L1'] == l1]
        rows_l = sub[['一级标签', '二级标签', '关键词']].values.tolist()
        nf = (len(rows_l) + 1000000 - 1) // 1000000
        for pi in range(nf):
            part = rows_l[pi*1000000:(pi+1)*1000000]
            name = f'{l1}_{DATE}_part{pi+1}.xlsx'
            ns = write_workbook(os.path.join(OUT_DIR, name), part, f'{l1}-{pi+1}')
            out_files.append((name, len(part), ns))
            print(f'📄 {name}: {len(part):,} / {ns} sheets')
    if merge:
        sub = df[df['L1'].isin(merge)]
        rows_l = sub[['一级标签', '二级标签', '关键词']].values.tolist()
        nf = (len(rows_l) + 1000000 - 1) // 1000000
        for pi in range(nf):
            part = rows_l[pi*1000000:(pi+1)*1000000]
            name = f'合并_{DATE}_part{pi+1}.xlsx'
            ns = write_workbook(os.path.join(OUT_DIR, name), part, f'合并-{pi+1}')
            out_files.append((name, len(part), ns))
            print(f'📄 {name}: {len(part):,} / {ns} sheets')

    print(f'\n✅ BGE 合并完成: {len(out_files)} 个文件, 总 {len(rows):,} 条')
    for name, n, ns in out_files:
        print(f'   - {name} ({n:,} 条, {ns} sheets)')


if __name__ == '__main__':
    main()
