"""元宝06合并文件（390.2万）分词分类 → 输出 D:\分词工具处理2.5\输出\20260807元宝06合并分类
输入: 8 个 part xlsx（3 列: 推广计划名称|推广单元名称|关键词名称）
输出: 按一级标签拆分（每 ≤50W，大标签拆 part，小标签合并）
"""
import sys, io, os, time, importlib.util, csv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from collections import Counter, defaultdict

ENGINE = r'D:\元宝分词分类引擎-V225\classify_engine_v225.py'
spec = importlib.util.spec_from_file_location("engine", ENGINE)
mod = importlib.util.module_from_spec(spec)
sys.modules['engine'] = mod
spec.loader.exec_module(mod)

import openpyxl
in_dir = r'D:\数据'
out_dir = r'D:\分词工具处理2.5\输出\20260807元宝06合并分类'
os.makedirs(out_dir, exist_ok=True)
MAX_PER_FILE = 500000

in_files = [f'baidu-TX元宝-06_合并_2026-08-07_去重过滤后_part{i}.xlsx' for i in range(1, 9)]

# ===== 第一步：分类 + 写临时 CSV =====
t0 = time.time()
total = 0
filtered = 0
l1_counter = Counter()
tmp_csv = os.path.join(out_dir, '_tmp_results.csv')

with open(tmp_csv, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['一级标签', '二级标签', '关键词'])
    for fname in in_files:
        fpath = os.path.join(in_dir, fname)
        wb = openpyxl.load_workbook(fpath, read_only=True)
        ws = wb.active
        cnt = 0
        t1 = time.time()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if not row or len(row) < 3 or not row[2]:
                continue
            kw = str(row[2]).strip()
            r = mod.classify_keyword(kw)
            if r:
                l1_counter[r['一级标签']] += 1
                writer.writerow([r['一级标签'], r['二级标签'], r['关键词']])
                total += 1
            else:
                filtered += 1
            cnt += 1
        wb.close()
        print(f'  {fname}: {cnt} 条, {time.time()-t1:.1f}s')

print(f'\n=== 分类完成 ===')
print(f'总有效: {total} | 过滤: {filtered} | 耗时: {time.time()-t0:.0f}s')
print(f'一级标签种类: {len(l1_counter)}')

with open(os.path.join(out_dir, '_统计报告.txt'), 'w', encoding='utf-8') as f:
    f.write(f'元宝06合并分类统计\n总有效: {total} | 过滤: {filtered}\n一级标签种类: {len(l1_counter)}\n\n')
    for l1, cnt in l1_counter.most_common():
        f.write(f'{cnt:7d} | {l1}\n')
print('统计已保存')

# ===== 第二步：规划输出文件 =====
big_labels = {l1: cnt for l1, cnt in l1_counter.items() if cnt > MAX_PER_FILE}
small_labels = [(l1, cnt) for l1, cnt in l1_counter.items() if cnt <= MAX_PER_FILE]
print(f'\n=== 输出规划 ===')
print(f'大标签（>50W）: {len(big_labels)} 个')
for l1, cnt in sorted(big_labels.items(), key=lambda x: -x[1]):
    n = (cnt + MAX_PER_FILE - 1) // MAX_PER_FILE
    print(f'  {cnt} → {n} 个文件 | {l1}')
print(f'小标签（≤50W）: {len(small_labels)} 个, 共 {sum(c for _, c in small_labels)} 条')

# ===== 第三步：写文件 =====
t0 = time.time()
label_rows = defaultdict(list)
with open(tmp_csv, 'r', encoding='utf-8', newline='') as f:
    reader = csv.reader(f)
    next(reader)
    for l1, l2, kw in reader:
        label_rows[l1].append((l2, kw))

for l1 in big_labels:
    safe = l1.replace('通用-', '').replace('行业-AI', 'AI').replace('-', '_').replace('/', '-')
    rows = label_rows[l1]
    n = (len(rows) + MAX_PER_FILE - 1) // MAX_PER_FILE
    for pi in range(1, n + 1):
        chunk = rows[(pi-1)*MAX_PER_FILE : pi*MAX_PER_FILE]
        outpath = os.path.join(out_dir, f'{safe}_part{pi}.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = l1[:31].replace('/', '-')
        ws.append(['一级标签', '二级标签', '关键词'])
        for l2, kw in chunk:
            ws.append([l1, l2, kw])
        wb.save(outpath)
        print(f'  写出: {outpath} ({len(chunk)} 条)')

small_items = sorted(small_labels, key=lambda x: -x[1])
merged_files = []
current = []
current_size = 0
file_idx = 1
for l1, cnt in small_items:
    if current_size + cnt > MAX_PER_FILE and current:
        merged_files.append((file_idx, current))
        current = []
        current_size = 0
        file_idx += 1
    current.append(l1)
    current_size += cnt
if current:
    merged_files.append((file_idx, current))

for file_idx, labels in merged_files:
    outpath = os.path.join(out_dir, f'合并小标签_{file_idx}.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '合并小标签'
    ws.append(['一级标签', '二级标签', '关键词'])
    for l1 in labels:
        for l2, kw in label_rows[l1]:
            ws.append([l1, l2, kw])
    wb.save(outpath)
    print(f'  写出: {outpath}')

os.remove(tmp_csv)
print(f'\n写出完成: 耗时 {time.time()-t0:.0f}s')
print(f'输出目录: {out_dir}')
