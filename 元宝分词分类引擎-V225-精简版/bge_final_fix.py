"""BGE 语义细分汇总（V188 修正版）：
1. 合并 part1~3 标签 CSV（L1 不新增，映射到已有 L1 的 L2）
2. 从原始 part1~3 过滤已细分词（保留重复行），重建"通用-其他"
用法: python bge_final_fix.py [base_dir] [prefix]（默认 20260803 全量）
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import defaultdict

# V197: 参数化（默认 20260803，可指定 20260807 元宝06）
base = sys.argv[1] if len(sys.argv) > 1 else r'D:\分词工具处理2.5\输出\20260803后合并分类'
prefix = sys.argv[2] if len(sys.argv) > 2 else '通用_其他_其他_part'
sub = os.path.join(base, '语义细分')

# ===== 1. 收集 part1~3 已细分词 =====
assigned_set = set()
for fname in os.listdir(sub):
    if fname.startswith(('part1_', 'part2_', 'part3_')) and 'remaining' not in fname and fname.endswith('.csv'):
        with open(os.path.join(sub, fname), encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split(',', 1)
                if parts:
                    assigned_set.add(parts[-1])
print(f'part1~3 已细分词: {len(assigned_set)}')

# ===== 2. 从 part 文件过滤（保留重复行）=====
MAX = 500000
all_remaining = []
part_exists = True
pi = 1
while part_exists:
    fpath = os.path.join(base, f'{prefix}{pi}.xlsx')
    if not os.path.exists(fpath):
        part_exists = False
        break
    wb = openpyxl.load_workbook(fpath, read_only=True)
    cnt = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 3 or not row[2]:
                continue
            kw = str(row[2]).strip()
            cnt += 1
            if kw not in assigned_set:
                all_remaining.append(kw)
    wb.close()
    print(f'  part{pi}: 原始 {cnt} 条')
    pi += 1

print(f'过滤后: {len(all_remaining)} 条')

# ===== 3. 重建 part1~n =====
for pi in range(1, 6):
    fpath = os.path.join(base, f'{prefix}{pi}.xlsx')
    if os.path.exists(fpath):
        os.remove(fpath)

n_parts = (len(all_remaining) + MAX - 1) // MAX
for pi in range(1, n_parts + 1):
    chunk = all_remaining[(pi-1)*MAX : pi*MAX]
    outpath = os.path.join(base, f'{prefix}{pi}.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '通用-其他'
    ws.append(['一级标签', '二级标签', '关键词'])
    for kw in chunk:
        ws.append([f'通用-其他-其他-26/08/08', '其他', kw])
    wb.save(outpath)
    print(f'  {outpath}: {len(chunk)} 条')

# ===== 4. 合并标签文件 =====
label_kws = defaultdict(list)
for fname in os.listdir(sub):
    if fname.startswith(('part1_', 'part2_', 'part3_')) and 'remaining' not in fname and fname.endswith('.csv'):
        with open(os.path.join(sub, fname), encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split(',', 1)
                if len(parts) == 2:
                    label_kws[parts[0]].append(parts[1])

print(f'\n=== 标签合并 ===')
total = 0
for label, kws_list in sorted(label_kws.items(), key=lambda x: -len(x[1])):
    print(f'  {len(kws_list):6d} | {label}')
    total += len(kws_list)
print(f'  合计: {total}')

for label, kws_list in label_kws.items():
    safe = label.replace('通用-', '').replace('行业-AI', 'AI').replace('-', '_')
    outpath = os.path.join(base, f'语义_{safe}.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = label[:31]
    ws.append(['一级标签', '二级标签', '关键词'])
    for kw in kws_list:
        ws.append([label, label.split('-')[-1], kw])
    wb.save(outpath)
print(f'\n完成！标签文件 {len(label_kws)} 个, 其他 part {n_parts} 个')
