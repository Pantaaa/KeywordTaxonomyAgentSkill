#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""元宝分词分类引擎 V225 — 分类 + 拆分输出工具

拆分规则（用户指定）：
  1. 每 100 万关键词 → 1 个文件
  2. 每个文件内每 50 万关键词 → 1 个 sheet
  3. 优先按标签合并拆分：
     - 数量 > 50 万的一级标签 → 单独 1 个文件（文件内按 50 万拆 sheet）
     - 数量 <= 50 万的一级标签 → 多个小标签合并成文件
  4. 严格遵守 100 万 / 50 万拆分逻辑

用法：
  python split_output.py                     # 全量处理待分词目录
  python split_output.py --limit 1000        # 仅处理前 N 条（验证用）
  python split_output.py --csv-only          # 只生成中间 CSV，不写 xlsx
"""
import sys, io, os, glob, argparse, time
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass
import pandas as pd
import openpyxl
from collections import Counter

ENGINE_DIR = os.path.dirname(os.path.abspath(__file__))
if ENGINE_DIR not in sys.path:
    sys.path.insert(0, ENGINE_DIR)
import classify_engine_v225 as engine

IN_DIR = r'D:\DSH-分词工作区\待分词'
OUT_DIR = r'D:\DSH-分词工作区\输出'
CSV_MID = os.path.join(OUT_DIR, '_分类结果_全量.csv')
DATE = time.strftime('%y-%m-%d')

FILE_CAP = 1_000_000   # 每文件最多关键词
SHEET_CAP = 500_000    # 每 sheet 最多关键词
L1_SOLO = 500_000      # 一级标签 >= 此值 → 单独文件（>50万）


def l1_of(label):
    """从引擎一级标签中提取 L1（品牌/竞品/行业-AI/通用）"""
    if not label:
        return '未知'
    if label.startswith('行业-AI'):
        return '行业-AI'
    return label.split('-')[0]


def classify_all(limit=None):
    """分类全部关键词 → 写中间 CSV（L1|二级标签|完整一级标签|关键词）
    用 openpyxl 流式读取，避免 pandas 全量加载，内存友好"""
    import openpyxl as _ox
    files = sorted(glob.glob(os.path.join(IN_DIR, '*.xlsx')))
    if not files:
        print('❌ 待分词目录无 xlsx 文件')
        return
    print(f'读取 {len(files)} 个文件，输出中间文件: {CSV_MID}')
    t0 = time.time()
    total = filtered = 0
    with open(CSV_MID, 'w', encoding='utf-8-sig', newline='') as fp:
        fp.write('一级标签,二级标签,关键词\n')
        for fi, f in enumerate(files, 1):
            print(f'[{fi}/{len(files)}] 处理 {os.path.basename(f)}')
            wb = _ox.load_workbook(f, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row and len(row) >= 3 and row[2]:
                    kw = str(row[2]).strip()
                    if not kw or kw == '关键词':
                        continue
                    r = engine.classify_keyword(kw)
                    if r:
                        fp.write(f'{r["一级标签"]},{r["二级标签"]},{r["关键词"]}\n')
                    else:
                        filtered += 1
                    total += 1
                    if limit and total >= limit:
                        break
                    if total % 200000 == 0:
                        print(f'  已处理 {total:,} 条，过滤 {filtered:,}，耗时 {time.time()-t0:.0f}s')
            wb.close()
            if limit and total >= limit:
                print(f'  达到验证上限 {limit}，停止')
                break
    print(f'✅ 分类完成: 有效 {total-filtered:,} | 过滤 {filtered:,} | 耗时 {time.time()-t0:.0f}s')


def load_mid():
    """读取中间 CSV → DataFrame"""
    df = pd.read_csv(CSV_MID, dtype=str, encoding='utf-8-sig')
    df['L1'] = df['一级标签'].map(l1_of)
    return df


def write_workbook(out_path, rows, sheet_prefix):
    """按 50 万/sheet 写入 xlsx（write_only 流式）"""
    wb = openpyxl.Workbook(write_only=True)
    n_sheets = (len(rows) + SHEET_CAP - 1) // SHEET_CAP
    for si in range(n_sheets):
        ws = wb.create_sheet(f'{sheet_prefix}-{si+1}')
        ws.append(['一级标签', '二级标签', '关键词'])
        chunk = rows[si * SHEET_CAP:(si + 1) * SHEET_CAP]
        for row in chunk:
            ws.append(row)
    wb.save(out_path)
    wb.close()
    return n_sheets


def split_output():
    """按规则拆分输出"""
    df = load_mid()
    total = len(df)
    print(f'中间数据: {total:,} 条')
    cnt = Counter(df['L1'])
    print('L1 分布:')
    for k, v in cnt.most_common():
        flag = '🔸单独文件' if v > L1_SOLO else '🔹合并文件'
        print(f'  {v:>8,} | {k:<6} {flag}')

    solo_l1 = [k for k, v in cnt.items() if v > L1_SOLO]
    merge_l1 = [k for k, v in cnt.items() if v <= L1_SOLO]
    out_files = []

    # 1) 大标签单独文件（文件内 100 万切分，sheet 50 万）
    for l1 in solo_l1:
        sub = df[df['L1'] == l1]
        rows = sub[['一级标签', '二级标签', '关键词']].values.tolist()
        n_files = (len(rows) + FILE_CAP - 1) // FILE_CAP
        for pi in range(n_files):
            part = rows[pi * FILE_CAP:(pi + 1) * FILE_CAP]
            name = f'{l1}_{DATE}_part{pi+1}.xlsx'
            path = os.path.join(OUT_DIR, name)
            ns = write_workbook(path, part, f'{l1}-{pi+1}')
            out_files.append((name, len(part), ns))
            print(f'  📄 {name}: {len(part):,} 条 / {ns} sheets')

    # 2) 小标签合并文件（按顺序拼接，100 万切文件、50 万切 sheet）
    if merge_l1:
        sub = df[df['L1'].isin(merge_l1)]
        rows = sub[['一级标签', '二级标签', '关键词']].values.tolist()
        n_files = (len(rows) + FILE_CAP - 1) // FILE_CAP
        for pi in range(n_files):
            part = rows[pi * FILE_CAP:(pi + 1) * FILE_CAP]
            name = f'合并_{DATE}_part{pi+1}.xlsx'
            path = os.path.join(OUT_DIR, name)
            ns = write_workbook(path, part, f'合并-{pi+1}')
            out_files.append((name, len(part), ns))
            print(f'  📄 {name}: {len(part):,} 条 / {ns} sheets')

    print(f'\n✅ 拆分完成: {len(out_files)} 个文件')
    for name, n, ns in out_files:
        print(f'   - {name} ({n:,} 条, {ns} sheets)')

    # 处理结束自动输出标签占比（用户要求：在反馈中体现，不单独输出文件）
    print('\n══════ 分类标签占比（处理结束反馈） ══════')
    print(f'总条数: {total:,}')
    print(f'\n【L1 一级标签占比】')
    for k, v in cnt.most_common():
        flag = '🔸单独文件' if v > L1_SOLO else '🔹合并文件'
        print(f'  {k:<8} {v:>9,} 条  {v/total*100:5.2f}%  {flag}')
    print(f'\n【"其他"占比（重点优化对象）】')
    df2 = df
    df2['L2'] = df2['二级标签'].str.replace(r'-\d+$', '', regex=True)
    # 完全未分类：通用-其他-其他（L1 与 L2 均为其他）
    oo_exact = df2[(df2['L1'] == '通用') & (df2['L2'] == '其他') & (df2['一级标签'].str.startswith('通用-其他-其他-'))]
    # L1 已分类但 L2 为其他（如 通用-政务-其他）
    oo_l1 = df2[(df2['L2'] == '其他') & ~df2['一级标签'].str.startswith('通用-其他-其他-')]
    other_cnt = df2[df2['L2'] == '其他']
    print(f'  🔴 完全未分类（通用-其他-其他）: {len(oo_exact):,} 条（{len(oo_exact)/total*100:.2f}%）')
    print(f'  🟡 L1已分类但L2为"其他": {len(oo_l1):,} 条（{len(oo_l1)/total*100:.2f}%）')
    print(f'  ⚪ 全部含"其他": {len(other_cnt):,} 条（{len(other_cnt)/total*100:.2f}%）')
    print(f'\n【L2 二级标签 Top 20】')
    l2c = Counter(df2['L2'])
    for k, v in l2c.most_common(20):
        print(f'  {k:<10} {v:>9,} 条  {v/total*100:5.2f}%')
    print('══════════════════════════════════════════')


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--limit', type=int, default=None, help='验证模式：只处理前 N 条')
    p.add_argument('--csv-only', action='store_true', help='只分类生成中间 CSV，不拆分')
    args = p.parse_args()

    if args.csv_only:
        classify_all(limit=args.limit)
        return
    if args.limit:
        classify_all(limit=args.limit)
        split_output()
    else:
        if not os.path.exists(CSV_MID):
            classify_all()
        split_output()


if __name__ == '__main__':
    main()
