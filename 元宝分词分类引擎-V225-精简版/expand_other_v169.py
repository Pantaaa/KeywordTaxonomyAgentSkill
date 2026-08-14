"""二级标签词根扩展工具 V169（独立运行）

用法：
  python expand_other_v169.py --input "分类结果.xlsx" --output "扩展结果.xlsx" [--min-freq 100]

规则：
  1. 仅对二级标签含"其他"的关键词进行扩展
  2. 词根提取：白名单词根（QUALITY_ROOTS）+ 关键词包含匹配
  3. 词根过滤：字数≥2、排除无意义词根（怎么/如何/什么/的话等）
  4. 扩展阈值：词根出现次数 ≥ min_freq（默认100）
  5. 归属规则：关键词命中多个词根时选最长（同长选高频）
  6. 未命中高频词根 → 保留"其他"
"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from collections import Counter, defaultdict

# ============ 一、无意义词根黑名单 ============
MEANINGLESS_ROOTS = set('''
怎么 如何 什么 怎样 为啥 为什么 哪个 哪些 可以 需要 这个 那个
有没有 是不是 能不能 做个 做一 写个 有吗 怎么办 怎么做 是什么
怎么样 好不好 多少 一个 一下 一些 还是 就是 不是 不会 不能
不要 不用 应该 可能 哪一 什么的 怎么用 怎么弄 怎么写 怎么读
怎么学 怎么练 怎么教 怎么画 怎么算 如何做 如何写 如何用 如何学
如何教 什么样 怎样的 哪样 咋样 咋弄 咋办 咋写 咋读 怎么的 如何的
的话 的了 的吗 的呢 的吧 这么 那么 做什么 是什么 干什么 咋整
关于 适合 有关 常见 我国 目前 相关 内容 方面 情况 问题 什么
'''.split())
for ch in '的了吗呢吧啊哦呀哈嗯么诶哟哎咳呵':
    MEANINGLESS_ROOTS.add(ch)

# ============ 二、高质量词根白名单（手工维护，可按需扩展） ============
QUALITY_ROOTS = [
    # 文本创作类
    '谈话记录', '谈心谈话', '经典十句', '文章精选', '写文章', '文章',
    '句子', '话语', '短句', '美句', '好句', '金句', '语录', '文案',
    '祝福', '祝福语', '问候', '感谢', '感恩', '感谢的话', '感谢老师',
    '生日快乐', '生日', '朋友圈', '发朋友圈', '励志', '正能量', '暖心',
    '唯美', '精选', '经典', '内容', '方法', '技巧', '教程', '模板',
    '大全', '下载', '软件', '官网', '价格', '攻略', '排名', '评价',
    # 教育类
    '简便运算', '扫一扫', '一对一', '作业', '答案', '计算', '练习',
    '老师', '家长', '孩子', '班主任', '毕业', '大学', '年级', '谈话',
    '谈心', '记录', '简便方法',
    # 体育类
    '看图写话', '放风筝', '舞蹈', '锻炼身体', '锻炼', '身体', '运动',
    '运动会', '高情商', '早上好', '教练', '跑步', '视频', '新闻',
    '分析报告', '报告', '分析', '比赛',
    # 其他
    '礼物', '健康', '美食', '开业', '激励', '温暖', '幽默',
    # 常见词根
    '电话', '微信', '地址', '电话号', '手机号', '联系', '咨询',
    '怎么样', '效果', '推荐', '介绍', '区别', '哪个好', '对比',
    '免费', '收费', '价格表', '多少钱', '费用', '行情',
    '证书', '考试', '报名', '时间', '地点', '流程', '条件', '要求',
    '作文', '范文', '开头', '结尾', '标题', '题目', '素材', '例文',
    '读后感', '观后感', '心得体会', '感悟', '感想',
    '成语', '诗词', '名言', '谚语', '歇后语', '对联',
    '方案', '计划', '总结', '报告', '策划', 'PPT', '表格',
    '朋友圈文案', '说说', '签名', '昵称', '头像',
    '早安', '晚安', '午安', '节日', '春节', '中秋', '端午', '国庆',
    '情人节', '母亲节', '父亲节', '教师节', '劳动节', '元旦',
    '减肥', '瘦身', '健身', '锻炼', '养生', '饮食', '食谱',
    '装修', '设计', '户型', '楼盘', '房价',
    '旅游', '景点', '攻略', '路线', '门票', '酒店',
    '股票', '基金', '理财', '投资', '保险',
    '招聘', '求职', '简历', '面试', '薪资',
    '电脑', '手机', '软件', 'APP', '应用', '程序',
    '游戏', '攻略', '充值', '账号', '礼包',
]

def is_chinese(s):
    return all('\u4e00' <= ch <= '\u9fff' for ch in s)

def is_meaningless(root):
    """判断词根是否无意义"""
    if root in MEANINGLESS_ROOTS:
        return True
    if root[0] in '的了吗呢吧啊哦呀':
        return True
    if root[-1] in '的了':
        return True
    for bad in ['的话', '的了', '的吗', '的呢', '什么的', '怎么弄', '怎么样']:
        if bad in root:
            return True
    return False

def extract_roots_from_kw(kw):
    """提取关键词中的白名单词根（最长优先，去重叠）"""
    found = [r for r in QUALITY_ROOTS if r in kw]
    result = []
    for r in sorted(found, key=lambda x: -len(x)):
        if not any(r in o and len(r) < len(o) for o in result):
            result.append(r)
    return result

def analyze_quality_roots(other_kws, min_freq=100):
    """统计白名单词根频率，返回 [(词根, 次数), ...]"""
    root_counter = Counter()
    for kw in other_kws:
        for r in extract_roots_from_kw(kw):
            root_counter[r] += 1
    return [(r, c) for r, c in root_counter.items() if c >= min_freq]

def assign_to_root(kw, roots):
    """分配关键词到词根：优先最长，同长优先高频"""
    best = None
    for root, cnt in roots:
        if root in kw:
            if best is None or len(root) > len(best[0]) or (len(root) == len(best[0]) and cnt > best[1]):
                best = (root, cnt)
    return best[0] if best else None

def expand_other(data, min_freq=100):
    """
    对分类结果扩展"其他"二级标签
    data: [(一级标签, 二级标签, 关键词), ...]
    返回: (扩展后列表, 统计信息)
    """
    others_by_l1 = defaultdict(list)
    kept = []
    for l1, l2, kw in data:
        if '其他' in l2:
            base = l1.rsplit('-', 1)[0]  # 去日期
            others_by_l1[base].append((l1, l2, kw))
        else:
            kept.append((l1, l2, kw))

    expanded = []
    stats = []
    # 每个 L1 组独立计数词根
    label_counter = Counter()
    for base, items in sorted(others_by_l1.items(), key=lambda x: -len(x[1])):
        other_kws = [kw for _, _, kw in items]
        roots = analyze_quality_roots(other_kws, min_freq)
        stats.append((base, len(items), roots))
        for l1, l2, kw in items:
            root = assign_to_root(kw, roots)
            if root:
                date_part = l1.rsplit('-', 1)[-1]
                # 二级标签单元号（按 L1+词根 独立计数）
                unit_key = (base, root)
                label_counter[unit_key] += 1
                unit = f'-{(label_counter[unit_key] - 1) // 5000 + 1}'
                new_l1 = f'{base}-{root}-{date_part}'
                new_l2 = f'{root}{unit}'
                expanded.append((new_l1, new_l2, kw))
            else:
                kept.append((l1, l2, kw))

    return kept + expanded, stats


# ============ 三、CLI 接口 ============
def main():
    args = sys.argv[1:]
    inp = out = None
    min_freq = 100
    if '--input' in args:
        inp = args[args.index('--input') + 1]
    if '--output' in args:
        out = args[args.index('--output') + 1]
    if '--min-freq' in args:
        min_freq = int(args[args.index('--min-freq') + 1])
    if not inp:
        print(__doc__)
        return

    # 读取
    import openpyxl
    wb = openpyxl.load_workbook(inp, read_only=True)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]: continue
        l1, l2 = str(row[0]), str(row[1]) if row[1] else ''
        if l1 == '一级标签':
            continue  # 跳过表头
        kw = str(row[2]) if len(row) > 2 and row[2] else ''
        data.append((l1, l2, kw))
    wb.close()
    print(f'输入: {len(data)} 条')

    # 扩展
    result, stats = expand_other(data, min_freq)
    before = sum(1 for _, l2, _ in data if '其他' in l2)
    after = sum(1 for _, l2, _ in result if '其他' in l2)
    print(f'扩展前"其他": {before}')
    print(f'扩展后"其他": {after}')
    print(f'减少: {before - after} ({(before-after)/before*100:.1f}%)')

    print(f'\n=== 词根扩展明细 ===')
    for base, total, roots in stats:
        if not roots:
            continue
        root_str = ', '.join(f'{r}({c})' for r, c in sorted(roots, key=lambda x: -x[1])[:10])
        print(f'  {base}: {total}条 → {root_str}')

    # 写出
    if out:
        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet()
        ws_out.append(['一级标签', '二级标签', '关键词'])
        for l1, l2, kw in result:
            ws_out.append([l1, l2, kw])
        wb_out.save(out)
        wb_out.close()
        print(f'\n输出: {out}')
        print(f'总条数: {len(result)}')

if __name__ == '__main__':
    main()
