"""最终输出：V172 全量分类 + 按一级标签拆分（每 50W，小标签合并）"""
import sys, io, os, time, importlib.util, csv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from collections import Counter, defaultdict

spec = importlib.util.spec_from_file_location("engine", r"D:\元宝分词分类引擎-V225\classify_engine_v225.py")
mod = importlib.util.module_from_spec(spec)
sys.modules['engine'] = mod
spec.loader.exec_module(mod)

target_dir = r'D:\分词工具处理2.5\待分词'
out_dir = r'D:\分词工具处理2.5\输出\20260803后合并分类'
os.makedirs(out_dir, exist_ok=True)
tmp_csv = os.path.join(out_dir, '_tmp_results.csv')

files = ['agent001.xlsx', 'agent002.xlsx', '0_guanjianci.xlsx', '1_guanjianci.xlsx',
         '5_guanjianci.xlsx', '4_guanjianci.xlsx', '3_guanjianci.xlsx', '2_guanjianci.xlsx',
         '8_guanjianci.xlsx', '7_guanjianci.xlsx', '6_guanjianci.xlsx']
import openpyxl

def read_keywords(fpath):
    wb = openpyxl.load_workbook(fpath, read_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            kw = str(row[0]).strip()
            if kw and kw != '关键词/营销要点' and kw != '关键词':
                yield kw
    wb.close()

# ===== 第一步：分类 + 写临时 CSV =====
t0 = time.time()
total = 0
filtered = 0
l1_counter = Counter()

with open(tmp_csv, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['一级标签', '二级标签', '关键词'])
    for fname in files:
        fpath = os.path.join(target_dir, fname)
        file_count = 0
        t1 = time.time()
        for kw in read_keywords(fpath):
            r = mod.classify_keyword(kw)
            if r:
                l1 = r['一级标签']
                l1_counter[l1] += 1
                writer.writerow([l1, r['二级标签'], r['关键词']])
                total += 1
            else:
                filtered += 1
            file_count += 1
        print(f'  {fname}: {file_count} 条, {time.time()-t1:.1f}s')

print(f'\n=== 分类完成 ===')
print(f'总有效: {total} | 过滤: {filtered} | 耗时: {time.time()-t0:.1f}s')
print(f'一级标签种类: {len(l1_counter)}')

# 保存分布
with open(os.path.join(out_dir, '_统计报告.txt'), 'w', encoding='utf-8') as f:
    f.write(f'V172 分类统计\n总有效: {total} | 过滤: {filtered}\n一级标签种类: {len(l1_counter)}\n\n')
    for l1, cnt in l1_counter.most_common():
        f.write(f'{cnt:7d} | {l1}\n')
print('统计已保存')

# ===== 第二步：规划输出文件 =====
MAX_PER_FILE = 500000
print(f'\n=== 输出规划 ===')
# 大标签：单独拆分成多个文件
big_labels = {l1: cnt for l1, cnt in l1_counter.items() if cnt > MAX_PER_FILE}
small_labels = {l1: cnt for l1, cnt in l1_counter.items() if cnt <= MAX_PER_FILE}
small_total = sum(small_labels.values())

print(f'大标签（>50W）: {len(big_labels)} 个')
for l1, cnt in sorted(big_labels.items(), key=lambda x: -x[1]):
    n_files = (cnt - 1) // MAX_PER_FILE + 1
    print(f'  {cnt:7d} → {n_files} 个文件 | {l1}')
print(f'小标签（≤50W）: {len(small_labels)} 个, 共 {small_total} 条')
if small_total > 0:
    n_small_files = (small_total - 1) // MAX_PER_FILE + 1
    print(f'  小标签合并 → {n_small_files} 个文件')

# ===== 第三步：按规划写出 xlsx =====
print(f'\n=== 写出文件 ===')
t2 = time.time()
# 大标签文件句柄
big_handles = {}  # l1 -> (wb, ws, count, part)
def get_big_writer(l1):
    if l1 not in big_handles:
        part = 1
        base = l1.rsplit('-', 1)[0].replace('-', '_')
        fpath = os.path.join(out_dir, f'{base}_part{part}.xlsx')
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.append(['一级标签', '二级标签', '关键词'])
        big_handles[l1] = {'wb': wb, 'ws': ws, 'count': 0, 'part': part, 'path': fpath}
    h = big_handles[l1]
    if h['count'] >= MAX_PER_FILE:
        # 关闭当前，开下一个
        h['wb'].save(h['path'])
        h['wb'].close()
        h['part'] += 1
        base = l1.rsplit('-', 1)[0].replace('-', '_')
        fpath = os.path.join(out_dir, f'{base}_part{h["part"]}.xlsx')
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.append(['一级标签', '二级标签', '关键词'])
        big_handles[l1] = {'wb': wb, 'ws': ws, 'count': 0, 'part': h['part'], 'path': fpath}
        h = big_handles[l1]
    return h

# 小标签合并写入（按 50W 拆）
small_writers = []
def get_small_writer():
    if not small_writers or small_writers[-1]['count'] >= MAX_PER_FILE:
        idx = len(small_writers) + 1
        fpath = os.path.join(out_dir, f'合并小标签_{idx}.xlsx')
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.append(['一级标签', '二级标签', '关键词'])
        small_writers.append({'wb': wb, 'ws': ws, 'count': 0, 'path': fpath})
    return small_writers[-1]

# 读 CSV 写出
with open(tmp_csv, 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    header = next(reader)  # 跳过表头
    for row in reader:
        if not row or len(row) < 3:
            continue
        l1, l2, kw = row[0], row[1], row[2]
        if l1 in big_labels:
            h = get_big_writer(l1)
            h['ws'].append([l1, l2, kw])
            h['count'] += 1
        else:
            h = get_small_writer()
            h['ws'].append([l1, l2, kw])
            h['count'] += 1

# 关闭所有
for l1, h in big_handles.items():
    h['wb'].save(h['path'])
    h['wb'].close()
for h in small_writers:
    h['wb'].save(h['path'])
    h['wb'].close()

# 删除临时 CSV
if os.path.exists(tmp_csv):
    os.remove(tmp_csv)

print(f'写出完成: {time.time()-t2:.1f}s')
print(f'\n=== 输出文件清单 ===')
for f in sorted(os.listdir(out_dir)):
    if f.endswith('.xlsx'):
        print(f'  {f}')
